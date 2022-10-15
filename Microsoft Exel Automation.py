from openpyxl import Workbook, load_workbook


print()
print()
print()


wb = load_workbook('Monthly Expenditure.xlsx')
ws = wb.active


print("Do You want to add any thing in your 'Money Expenditure'?")
print()
query = input("> ")


def purchased_date():
    print("When did you purchase it?")
    date_of_purchase = input("> ")
    ws['A7'].value = date_of_purchase
    wb.save("Monthly Expenditure.xlsx")

def Money_before_spending():
    print("What was your balance?")
    print(ws["M5"].value, "correct?")
    Balance_Yes_or_No = input("> ")
    if "s" in Balance_Yes_or_No:
        ws['C7'].value = ws["M5"].value
        wb.save("Monthly Expenditure.xlsx")
    elif "n" in Balance_Yes_or_No:
        print("What is the balance?")
        balance = input("> ₹")
        ws['C7'].value = balance
        wb.save("Monthly Expenditure.xlsx")

def Item_purchased():
    print("What item/items was/were purchased?")
    Item_purchased = input("> ")
    ws['E7'].value = Item_purchased
    wb.save("Monthly Expenditure.xlsx")
    
def Quantity___Cost_of_one_item___Total_cost():


    print("What was/were the quantity/Quantities?")
    Quantity = int(input("> "))
    ws['G7'].value = Quantity
    wb.save("Monthly Expenditure.xlsx")



    print("What was the Cost of one item?")
    Cost_of_one_item = int(input("> "))
    ws['I7'].value = Cost_of_one_item
    wb.save("Monthly Expenditure.xlsx")



    print("What was the total cost of the item/items?")
    Total_cost_auto = Cost_of_one_item * Quantity
    print(Total_cost_auto, "Correct?")
    Total_cost_Yes_or_No = input("> ")
    if "s" in Total_cost_Yes_or_No:
        ws['k7'].value = Total_cost_auto
        wb.save("Monthly Expenditure.xlsx")
    elif "n" in Total_cost_Yes_or_No:
        print("What is the total money?")
        Total_cost_Manual = input("> ")
        ws['k7'].value = Total_cost_Manual
        wb.save("Monthly Expenditure.xlsx")

def Money_left_after_purchasing():
    print("What was the money left with you?")
    Money_Left = input("> ")
    ws['M7'].value = Money_Left
    wb.save("Monthly Expenditure.xlsx")


if "s" in query:

    purchased_date()
    Money_before_spending()
    Item_purchased()
    Quantity___Cost_of_one_item___Total_cost()
    Money_left_after_purchasing()

else:
    quit()
